"""
My Library — a small local web app for cataloguing a personal book collection.

Run with:  python3 app.py
Then open  http://localhost:5000   on this PC, or
           http://<this-PC's-LAN-IP>:5000   from another device on the same WiFi.
"""

import base64
import json
import mimetypes
import os
import re
import sqlite3
import subprocess
import sys
import tempfile
import time
import uuid
from collections import namedtuple
from pathlib import Path
import csv
import io
from datetime import date

import requests
from flask import Flask, g, jsonify, request, send_file, render_template, abort, Response
from PIL import Image, ImageOps
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage
from hijridate import Gregorian

from options import OPTION_FIELDS

BASE_DIR = Path(__file__).parent.resolve()
DATA_DIR = BASE_DIR / "data"
COVERS_DIR = DATA_DIR / "covers"
DB_PATH = DATA_DIR / "library.db"
CONFIG_PATH = DATA_DIR / "config.json"
CJPEG_PATH = BASE_DIR / "tools" / "mozjpeg" / ("cjpeg.exe" if sys.platform == "win32" else "cjpeg")

DATA_DIR.mkdir(parents=True, exist_ok=True)
COVERS_DIR.mkdir(parents=True, exist_ok=True)

app = Flask(__name__)

# ----------------------------------------------------------------------------
# Database
# ----------------------------------------------------------------------------

def get_db():
    if "db" not in g:
        g.db = sqlite3.connect(str(DB_PATH))
        g.db.row_factory = sqlite3.Row
    return g.db


@app.teardown_appcontext
def close_db(exception=None):
    db = g.pop("db", None)
    if db is not None:
        db.close()


def init_db():
    conn = sqlite3.connect(str(DB_PATH))
    conn.execute("""
        CREATE TABLE IF NOT EXISTS books (
            id TEXT PRIMARY KEY,
            accession INTEGER,
            title TEXT NOT NULL,
            author TEXT,
            death_year TEXT,
            publisher TEXT,
            edition TEXT,
            year TEXT,
            genre TEXT,
            shelf TEXT,
            notes TEXT,
            cover_path TEXT,
            pdf_path TEXT,
            added_at INTEGER,
            language TEXT,
            is_translation TEXT,
            copy_type TEXT,
            volume TEXT,
            shelf_position TEXT,
            shelf_side TEXT
        )
    """)
    for col in ("language", "is_translation", "copy_type", "volume", "shelf_position", "shelf_side", "set_group_id", "translator"):
        try:
            conn.execute(f"ALTER TABLE books ADD COLUMN {col} TEXT")
        except sqlite3.OperationalError:
            pass  # column already exists, safe to ignore
    conn.commit()
    conn.close()


init_db()

# ----------------------------------------------------------------------------
# Book field definitions — single place mapping API/CSV keys to DB columns.
# Add a row here (plus an ALTER TABLE above) when introducing a genuinely new
# column; this drives row_to_dict, create/update, and CSV export/import.
# ----------------------------------------------------------------------------

Field = namedtuple("Field", ["json_key", "db_col", "digits_only"])

BOOK_FIELDS = [
    Field("title", "title", False),
    Field("author", "author", False),
    Field("deathYear", "death_year", False),
    Field("publisher", "publisher", False),
    Field("edition", "edition", False),
    Field("year", "year", False),
    Field("volume", "volume", True),
    Field("genre", "genre", False),
    Field("language", "language", False),
    Field("isTranslation", "is_translation", False),
    Field("translator", "translator", False),
    Field("copyType", "copy_type", False),
    Field("shelf", "shelf", False),
    Field("shelfPosition", "shelf_position", False),
    Field("shelfSide", "shelf_side", False),
    Field("notes", "notes", False),
    Field("setGroupId", "set_group_id", False),
]

# ----------------------------------------------------------------------------
# Config (pdf folder location, etc.)
# ----------------------------------------------------------------------------

def load_config():
    defaults = {"pdf_folder": "", "gemini_api_key": ""}
    if CONFIG_PATH.exists():
        try:
            return {**defaults, **json.loads(CONFIG_PATH.read_text())}
        except Exception:
            pass
    return defaults


def save_config(cfg):
    CONFIG_PATH.write_text(json.dumps(cfg, indent=2))


# ----------------------------------------------------------------------------
# Helpers
# ----------------------------------------------------------------------------

def row_to_dict(row):
    d = dict(row)
    out = {f.json_key: d[f.db_col] or "" for f in BOOK_FIELDS}
    out["id"] = d["id"]
    out["accession"] = d["accession"]
    out["hasCover"] = bool(d.get("cover_path"))
    out["hasPdf"] = bool(d.get("pdf_path")) and Path(d["pdf_path"]).exists()
    out["addedAt"] = d["added_at"]
    return out

def strip_arabic_diacritics(s):
    # removes tashkeel/harakat so "الْقُرْطُبِيّ" and "القرطبي" match the same
    return re.sub(r"[\u0610-\u061A\u064B-\u065F\u06D6-\u06DC\u06DF-\u06E8\u06EA-\u06ED]", "", s or "")

def normalize(s):
    if not s:
        return ""
    s = strip_arabic_diacritics(s)
    s = s.lower().strip()
    s = re.sub(r"[^\w\s]", "", s)
    s = re.sub(r"\s+", " ", s)
    return s

def guess_ext(mime, url=""):
    ext = mimetypes.guess_extension(mime) if mime else None
    if ext:
        return ext.replace(".jpe", ".jpg")
    m = re.search(r"\.(jpg|jpeg|png|webp|gif)(\?|$)", url, re.I)
    if m:
        return "." + m.group(1).lower()
    return ".jpg"


def compress_image_bytes(data, max_dim=1000, quality=85, cjpeg_path=CJPEG_PATH):
    """Resize to max_dim on the longest edge and re-encode with mozJPEG's
    cjpeg. Covers are only ever displayed at well under 150px in the UI, so
    this cuts multi-MB phone-camera photos down to a few hundred KB with no
    visible quality loss."""
    if not cjpeg_path.exists():
        raise FileNotFoundError(f"mozjpeg cjpeg binary not found at {cjpeg_path}")

    img = Image.open(io.BytesIO(data))
    img = ImageOps.exif_transpose(img)  # respect camera rotation before resizing
    if img.mode != "RGB":
        img = img.convert("RGB")  # drop alpha/CMYK, JPEG has no transparency

    w, h = img.size
    longest = max(w, h)
    if longest > max_dim:
        scale = max_dim / longest
        img = img.resize((round(w * scale), round(h * scale)), Image.LANCZOS)

    with tempfile.NamedTemporaryFile(suffix=".ppm", delete=False) as tmp:
        img.save(tmp, format="PPM")
        tmp_path = Path(tmp.name)

    try:
        result = subprocess.run(
            [str(cjpeg_path), "-quality", str(quality), "-optimize", "-progressive", str(tmp_path)],
            capture_output=True, check=True,
        )
        return result.stdout  # encoded JPEG bytes; no EXIF/GPS carried over
    finally:
        tmp_path.unlink(missing_ok=True)


def save_cover_from_bytes(book_id, raw, mime=None):
    for f in COVERS_DIR.glob(f"{book_id}.*"):
        f.unlink(missing_ok=True)
    try:
        compressed = compress_image_bytes(raw)
        path = COVERS_DIR / f"{book_id}.jpg"
        path.write_bytes(compressed)
    except Exception as e:
        print(f"[cover] mozjpeg compression failed ({e}), saving original uncompressed")
        path = COVERS_DIR / f"{book_id}{guess_ext(mime)}"
        path.write_bytes(raw)
    return str(path)


def save_cover_from_base64(book_id, data_b64, mime):
    return save_cover_from_bytes(book_id, base64.b64decode(data_b64), mime)


def delete_cover(book_id):
    for f in COVERS_DIR.glob(f"{book_id}.*"):
        f.unlink(missing_ok=True)


def field_value(field, data):
    value = (data.get(field.json_key) or "").strip()
    if field.digits_only:
        value = re.sub(r"\D", "", value)
    return value


# ----------------------------------------------------------------------------
# Pages
# ----------------------------------------------------------------------------

def today_dates():
    today = date.today()
    greg_str = f"{today.day} {today.strftime('%B %Y')}"
    hijri = Gregorian(today.year, today.month, today.day).to_hijri()
    hijri_str = f"{hijri.day} {hijri.month_name()} {hijri.year} AH"
    return {"gregorian": greg_str, "hijri": hijri_str}


@app.route("/")
def index():
    return render_template("index.html", today=today_dates())


# ----------------------------------------------------------------------------
# Book CRUD
# ----------------------------------------------------------------------------

@app.route("/api/books", methods=["GET"])
def list_books():
    db = get_db()
    rows = db.execute("SELECT * FROM books ORDER BY title COLLATE NOCASE").fetchall()
    return jsonify([row_to_dict(r) for r in rows])


@app.route("/api/books", methods=["POST"])
def create_book():
    data = request.get_json(force=True) or {}
    title = (data.get("title") or "").strip()
    if not title:
        return jsonify({"error": "Title is required."}), 400

    db = get_db()
    book_id = uuid.uuid4().hex
    max_acc = db.execute("SELECT MAX(accession) AS m FROM books").fetchone()["m"] or 0

    cover_path = None
    cover = data.get("cover")
    if cover and cover.get("data"):
        cover_path = save_cover_from_base64(book_id, cover["data"], cover.get("mime"))

    columns = ["id", "accession"] + [f.db_col for f in BOOK_FIELDS] + ["cover_path", "pdf_path", "added_at"]
    values = [book_id, max_acc + 1] + [field_value(f, data) for f in BOOK_FIELDS] + [cover_path, None, int(time.time())]
    db.execute(
        f"INSERT INTO books ({', '.join(columns)}) VALUES ({', '.join('?' * len(values))})",
        values,
    )
    db.commit()
    row = db.execute("SELECT * FROM books WHERE id = ?", (book_id,)).fetchone()
    return jsonify(row_to_dict(row)), 201


@app.route("/api/books/<book_id>", methods=["PUT"])
def update_book(book_id):
    db = get_db()
    existing = db.execute("SELECT * FROM books WHERE id = ?", (book_id,)).fetchone()
    if not existing:
        return jsonify({"error": "Book not found."}), 404

    data = request.get_json(force=True) or {}
    if "title" in data and not (data["title"] or "").strip():
        return jsonify({"error": "Title cannot be empty."}), 400

    updates, params = [], []
    for f in BOOK_FIELDS:
        if f.json_key in data:
            updates.append(f"{f.db_col} = ?")
            params.append((data[f.json_key] or "").strip())

    if data.get("removeCover"):
        delete_cover(book_id)
        updates.append("cover_path = ?")
        params.append(None)
    elif data.get("cover") and data["cover"].get("data"):
        cover_path = save_cover_from_base64(book_id, data["cover"]["data"], data["cover"].get("mime"))
        updates.append("cover_path = ?")
        params.append(cover_path)

    if updates:
        params.append(book_id)
        db.execute(f"UPDATE books SET {', '.join(updates)} WHERE id = ?", params)
        db.commit()

    row = db.execute("SELECT * FROM books WHERE id = ?", (book_id,)).fetchone()
    return jsonify(row_to_dict(row))


@app.route("/api/books/<book_id>", methods=["DELETE"])
def delete_book(book_id):
    db = get_db()
    row = db.execute("SELECT * FROM books WHERE id = ?", (book_id,)).fetchone()
    if not row:
        return jsonify({"error": "Book not found."}), 404
    delete_cover(book_id)
    db.execute("DELETE FROM books WHERE id = ?", (book_id,))
    db.commit()
    return jsonify({"ok": True})


@app.route("/api/books/bulk_delete", methods=["POST"])
def bulk_delete_books():
    data = request.get_json(force=True) or {}
    ids = data.get("ids") or []
    if not isinstance(ids, list) or not ids:
        return jsonify({"error": "No book ids provided."}), 400
    db = get_db()
    deleted = []
    for book_id in ids:
        if db.execute("SELECT id FROM books WHERE id = ?", (book_id,)).fetchone():
            delete_cover(book_id)
            db.execute("DELETE FROM books WHERE id = ?", (book_id,))
            deleted.append(book_id)
    db.commit()
    return jsonify({"deleted": deleted, "notFound": [i for i in ids if i not in deleted]})


BULK_UPDATE_FIELDS = [f for f in BOOK_FIELDS if f.json_key in ("shelf", "shelfPosition", "shelfSide")]


@app.route("/api/books/bulk_update", methods=["POST"])
def bulk_update_books():
    data = request.get_json(force=True) or {}
    ids = data.get("ids") or []
    if not isinstance(ids, list) or not ids:
        return jsonify({"error": "No book ids provided."}), 400

    fields_to_set = [f for f in BULK_UPDATE_FIELDS if f.json_key in data]
    if not fields_to_set:
        return jsonify({"error": "No fields to update."}), 400

    db = get_db()
    set_clause = ", ".join(f"{f.db_col} = ?" for f in fields_to_set)
    base_params = [field_value(f, data) for f in fields_to_set]

    updated = []
    for book_id in ids:
        if db.execute("SELECT id FROM books WHERE id = ?", (book_id,)).fetchone():
            db.execute(f"UPDATE books SET {set_clause} WHERE id = ?", base_params + [book_id])
            updated.append(book_id)
    db.commit()
    return jsonify({"updated": updated, "notFound": [i for i in ids if i not in updated]})


@app.route("/api/books/<book_id>/cover", methods=["GET"])
def get_cover(book_id):
    matches = list(COVERS_DIR.glob(f"{book_id}.*"))
    if not matches:
        abort(404)
    return send_file(matches[0])


@app.route("/api/books/<book_id>/pdf", methods=["GET"])
def get_pdf(book_id):
    db = get_db()
    row = db.execute("SELECT pdf_path FROM books WHERE id = ?", (book_id,)).fetchone()
    if not row or not row["pdf_path"] or not Path(row["pdf_path"]).exists():
        abort(404)
    return send_file(row["pdf_path"], mimetype="application/pdf")


# ----------------------------------------------------------------------------
# Dropdown options (genre/publisher/language/etc. — see options.py)
# ----------------------------------------------------------------------------

@app.route("/api/options", methods=["GET"])
def get_options():
    return jsonify(OPTION_FIELDS)


# ----------------------------------------------------------------------------
# Settings
# ----------------------------------------------------------------------------

@app.route("/api/settings", methods=["GET"])
def get_settings():
    cfg = load_config()
    return jsonify({"pdf_folder": cfg.get("pdf_folder", ""), "hasGeminiKey": bool(cfg.get("gemini_api_key"))})


@app.route("/api/settings", methods=["POST"])
def set_settings():
    data = request.get_json(force=True) or {}
    cfg = load_config()
    if "pdf_folder" in data:
        cfg["pdf_folder"] = (data["pdf_folder"] or "").strip()
    if data.get("clearGeminiKey"):
        cfg["gemini_api_key"] = ""
    elif data.get("gemini_api_key"):
        cfg["gemini_api_key"] = data["gemini_api_key"].strip()
    save_config(cfg)
    return jsonify({"pdf_folder": cfg.get("pdf_folder", ""), "hasGeminiKey": bool(cfg.get("gemini_api_key"))})


# ----------------------------------------------------------------------------
# Cover info extraction (Gemini)
# ----------------------------------------------------------------------------

class GeminiConfigError(Exception):
    pass


class GeminiNetworkError(Exception):
    pass


class GeminiResponseError(Exception):
    pass


GEMINI_MODEL = "gemini-flash-latest"  # alias Google keeps pointed at a current flash model — avoids
# hardcoding a dated model id (e.g. gemini-2.5-flash) that gets deprecated for new API keys over time
GEMINI_ENDPOINT = f"https://generativelanguage.googleapis.com/v1beta/models/{GEMINI_MODEL}:generateContent"

EXTRACTION_PROMPT = """You are looking at a photo of a book's cover or title page. Read exactly what is printed on it and return ONLY a JSON object (no other text) with these keys:
- "title": the book's title as printed.
- "author": the author's name as printed.
- "deathYear": the author's death year exactly as printed, often in parentheses after their name (may include "AH", "CE", or the Arabic هـ). Leave "" if not shown.
- "publisher": the publisher's name as printed.
- "translator": the translator's name, ONLY if the cover credits someone as a translator (this means the book is a translation). Leave "" otherwise.
- "isTranslation": "Translation" if a translator or translation is indicated on the cover, otherwise "Original".
- "notABookCover": true if this photo is not a book cover or title page at all (e.g. a blank page, a random object, a photo of a person), otherwise false.

Leave any field "" if it is not visible or you are not confident about it. Do not guess or invent values. Respond with raw JSON only, no markdown code fences."""


def prepare_image_for_gemini(raw, mime=None, max_dim=1568):
    """Downscale/normalize a cover photo before sending it to Gemini. Kept separate
    from compress_image_bytes()/mozjpeg since this image is never written to disk —
    it's sent to the API and discarded."""
    img = Image.open(io.BytesIO(raw))
    img = ImageOps.exif_transpose(img)
    if img.mode != "RGB":
        img = img.convert("RGB")
    w, h = img.size
    longest = max(w, h)
    if longest > max_dim:
        scale = max_dim / longest
        img = img.resize((round(w * scale), round(h * scale)), Image.LANCZOS)
    buf = io.BytesIO()
    img.save(buf, format="JPEG", quality=92)
    return buf.getvalue(), "image/jpeg"


def call_gemini_extract(image_bytes, mime, api_key):
    payload = {
        "contents": [{"parts": [
            {"text": EXTRACTION_PROMPT},
            {"inline_data": {"mime_type": mime, "data": base64.b64encode(image_bytes).decode("ascii")}},
        ]}],
        "generationConfig": {"response_mime_type": "application/json"},
    }
    try:
        resp = requests.post(GEMINI_ENDPOINT, params={"key": api_key}, json=payload, timeout=30)
    except requests.RequestException as e:
        raise GeminiNetworkError(f"Could not reach Gemini: {e}")

    if resp.status_code in (400, 403):
        raise GeminiConfigError("Gemini rejected the request — check that your API key is valid.")
    if resp.status_code == 429:
        raise GeminiNetworkError("Gemini rate limit reached — try again shortly.")
    if not resp.ok:
        raise GeminiNetworkError(f"Gemini request failed (HTTP {resp.status_code}).")

    try:
        text = resp.json()["candidates"][0]["content"]["parts"][0]["text"]
        parsed = json.loads(text)
    except Exception:
        raise GeminiResponseError("Gemini returned an unexpected response. Please try again or enter details manually.")

    return {
        "title": str(parsed.get("title") or "").strip(),
        "author": str(parsed.get("author") or "").strip(),
        "deathYear": str(parsed.get("deathYear") or "").strip(),
        "publisher": str(parsed.get("publisher") or "").strip(),
        "translator": str(parsed.get("translator") or "").strip(),
        "isTranslation": parsed.get("isTranslation") if parsed.get("isTranslation") in ("Original", "Translation") else "",
        "notABookCover": bool(parsed.get("notABookCover")),
    }


@app.route("/api/extract_cover", methods=["POST"])
def extract_cover():
    data = request.get_json(force=True) or {}
    cover = data.get("cover") or {}
    b64 = cover.get("data")
    if not b64:
        return jsonify({"error": "No cover image provided."}), 400

    api_key = load_config().get("gemini_api_key", "")
    if not api_key:
        return jsonify({"error": "Gemini API key not set. Add one under AI Settings first."}), 400

    try:
        raw = base64.b64decode(b64)
    except Exception:
        return jsonify({"error": "Could not decode the image data."}), 400

    try:
        img_bytes, img_mime = prepare_image_for_gemini(raw, cover.get("mime"))
    except Exception:
        return jsonify({"error": "Could not process the image."}), 400

    try:
        result = call_gemini_extract(img_bytes, img_mime, api_key)
    except GeminiConfigError as e:
        return jsonify({"error": str(e)}), 400
    except (GeminiNetworkError, GeminiResponseError) as e:
        return jsonify({"error": str(e)}), 502

    return jsonify(result)


# ----------------------------------------------------------------------------
# PDF folder scan
# ----------------------------------------------------------------------------

@app.route("/api/scan_pdfs", methods=["POST"])
def scan_pdfs():
    cfg = load_config()
    folder = cfg.get("pdf_folder", "")
    if not folder or not Path(folder).is_dir():
        return jsonify({"error": "PDF folder is not set or does not exist."}), 400

    db = get_db()
    candidates = db.execute(
        "SELECT * FROM books WHERE pdf_path IS NULL OR pdf_path = ''"
    ).fetchall()
    already_linked_paths = {
        r["pdf_path"] for r in db.execute(
            "SELECT pdf_path FROM books WHERE pdf_path IS NOT NULL AND pdf_path != ''"
        ).fetchall()
    }

    linked, skipped_linked, skipped_unmatched = [], [], []
    candidates = list(candidates)

    for root, _dirs, files in os.walk(folder):
        for fname in files:
            if not fname.lower().endswith(".pdf"):
                continue
            full_path = str(Path(root) / fname)

            if full_path in already_linked_paths:
                skipped_linked.append(fname)
                continue

            stem = fname[:-4]
            parts = [p.strip() for p in stem.split(" - ")]
            f_title = parts[0] if len(parts) >= 1 else ""
            f_author = parts[1] if len(parts) >= 2 else ""
            f_publisher = parts[2] if len(parts) >= 3 else ""
            f_year = parts[3] if len(parts) >= 4 else ""

            if not f_title:
                skipped_unmatched.append(fname)
                continue

            pool = [c for c in candidates if normalize(c["title"]) == normalize(f_title)]
            if f_author:
                narrowed = [c for c in pool if normalize(f_author) in normalize(c["author"] or "")]
                if narrowed:
                    pool = narrowed
            if f_publisher:
                narrowed = [c for c in pool if normalize(f_publisher) in normalize(c["publisher"] or "")]
                if narrowed:
                    pool = narrowed
            if f_year:
                narrowed = [c for c in pool if normalize(f_year) == normalize(c["year"] or "")]
                if narrowed:
                    pool = narrowed

            if pool:
                chosen = pool[0]
                db.execute("UPDATE books SET pdf_path = ? WHERE id = ?", (full_path, chosen["id"]))
                candidates = [c for c in candidates if c["id"] != chosen["id"]]
                linked.append({"file": fname, "matchedTitle": chosen["title"]})
            else:
                skipped_unmatched.append(fname)

    db.commit()
    return jsonify({
        "linked": linked,
        "skippedAlreadyLinked": skipped_linked,
        "skippedUnmatched": skipped_unmatched,
    })

CSV_FIELDS = ["id"] + [f.json_key for f in BOOK_FIELDS] + ["pdfPath"]

# Cover images cannot be embedded in the plain-text CSV in any visual sense,
# so CSV export/import stays text-only. The XLSX export/import (below) is the
# format that carries actual cover pictures, for opening in Excel.

@app.route("/api/export/csv", methods=["GET"])
def export_csv():
    db = get_db()
    rows = db.execute("SELECT * FROM books ORDER BY title COLLATE NOCASE").fetchall()
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=CSV_FIELDS)
    writer.writeheader()
    for r in rows:
        csv_row = {f.json_key: r[f.db_col] or "" for f in BOOK_FIELDS}
        csv_row["id"] = r["id"]
        csv_row["pdfPath"] = r["pdf_path"] or ""
        writer.writerow(csv_row)
    return Response(
        "\ufeff" + output.getvalue(),
        mimetype="text/csv; charset=utf-8",
        headers={"Content-Disposition": "attachment; filename=my-library-export.csv"},
    )


@app.route("/api/import/csv", methods=["POST"])
def import_csv():
    file = request.files.get("file")
    if not file:
        return jsonify({"error": "No file uploaded."}), 400
    try:
        text = file.stream.read().decode("utf-8-sig")
    except Exception:
        return jsonify({"error": "Could not read the file as UTF-8 text."}), 400

    reader = csv.DictReader(io.StringIO(text))
    db = get_db()
    result = process_import_rows(db, reader)
    return jsonify(result)


# ----------------------------------------------------------------------------
# Shared import logic (used by both CSV and XLSX import) \u2014 a "row" here is any
# dict-like object keyed by the same json_keys as BOOK_FIELDS, plus optionally
# "id", "pdfPath", and "_coverBytes" (raw image bytes, or absent/None to leave
# the cover untouched). CSV rows never carry "_coverBytes" \u2014 CSV is plain text
# and never touches covers on import.
# ----------------------------------------------------------------------------

def apply_import_row(db, row_id, values, pdf_path, cover_bytes):
    existing = db.execute("SELECT id FROM books WHERE id = ?", (row_id,)).fetchone() if row_id else None

    if existing:
        set_clause = ", ".join(f"{col} = ?" for col in values) + ", pdf_path = ?"
        params = [*values.values(), pdf_path]
        if cover_bytes is not None:
            cover_path = save_cover_from_bytes(row_id, cover_bytes)
            set_clause += ", cover_path = ?"
            params.append(cover_path)
        params.append(row_id)
        db.execute(f"UPDATE books SET {set_clause} WHERE id = ?", params)
        return "updated"

    new_id = uuid.uuid4().hex
    max_acc = db.execute("SELECT MAX(accession) AS m FROM books").fetchone()["m"] or 0
    cover_path = None
    if cover_bytes is not None:
        cover_path = save_cover_from_bytes(new_id, cover_bytes)
    columns = ["id", "accession"] + list(values.keys()) + ["cover_path", "pdf_path", "added_at"]
    insert_values = [new_id, max_acc + 1] + list(values.values()) + [cover_path, pdf_path, int(time.time())]
    db.execute(
        f"INSERT INTO books ({', '.join(columns)}) VALUES ({', '.join('?' * len(insert_values))})",
        insert_values,
    )
    return "created"


def process_import_rows(db, rows):
    created = updated = skipped = 0
    for row in rows:
        title = (row.get("title") or "").strip()
        if not title:
            skipped += 1
            continue

        row_id = (row.get("id") or "").strip()
        pdf_path = (row.get("pdfPath") or "").strip() or None
        values = {f.db_col: field_value(f, row) for f in BOOK_FIELDS}
        cover_bytes = row.get("_coverBytes")

        outcome = apply_import_row(db, row_id, values, pdf_path, cover_bytes)
        if outcome == "updated":
            updated += 1
        else:
            created += 1

    db.commit()
    return {"created": created, "updated": updated, "skipped": skipped}


# ----------------------------------------------------------------------------
# XLSX export/import \u2014 an actual Excel workbook with a real, viewable cover
# picture per row (a plain CSV can only ever hold text, never an image).
#
# The cover is embedded as an actual picture (full original bytes \u2014 only its
# on-screen display size is shrunk, not the embedded data), not as base64 text
# in a cell: Excel/openpyxl cells are hard-capped at 32,767 characters, and a
# compressed cover's base64 form routinely exceeds that, which silently
# truncates and corrupts it. Embedding it as a real image sidesteps that
# limit entirely and is also what makes it visible as a picture in Excel.
# Import reads the cover back directly from the embedded picture (matched to
# its row via the image's anchor), not from any text column.
# ----------------------------------------------------------------------------

XLSX_TEXT_FIELDS = ["id"] + [f.json_key for f in BOOK_FIELDS] + ["pdfPath"]
XLSX_FORCE_TEXT_COLS = {"id", "year", "deathYear", "volume", "pdfPath"}
COVER_DISPLAY_MAX_SIZE = (90, 120)  # px \u2014 on-screen size only; embedded bytes stay full quality


@app.route("/api/export/xlsx", methods=["GET"])
def export_xlsx():
    db = get_db()
    rows = db.execute("SELECT * FROM books ORDER BY title COLLATE NOCASE").fetchall()

    header = ["Cover"] + XLSX_TEXT_FIELDS
    wb = Workbook()
    ws = wb.active
    ws.title = "Library"
    ws.append(header)
    ws.column_dimensions["A"].width = 13

    for r_idx, r in enumerate(rows, start=2):
        csv_row = {f.json_key: r[f.db_col] or "" for f in BOOK_FIELDS}
        csv_row["id"] = r["id"]
        csv_row["pdfPath"] = r["pdf_path"] or ""

        ws.append([""] + [csv_row[col] for col in XLSX_TEXT_FIELDS])

        for col_idx, col_name in enumerate(header, start=1):
            if col_name in XLSX_FORCE_TEXT_COLS:
                ws.cell(row=r_idx, column=col_idx).number_format = "@"

        cover_path = r["cover_path"]
        if cover_path and Path(cover_path).exists():
            try:
                raw_bytes = Path(cover_path).read_bytes()
                with Image.open(io.BytesIO(raw_bytes)) as probe:
                    w, h = probe.size
                scale = min(COVER_DISPLAY_MAX_SIZE[0] / w, COVER_DISPLAY_MAX_SIZE[1] / h, 1)
                disp_w, disp_h = max(1, round(w * scale)), max(1, round(h * scale))
                xl_img = XLImage(io.BytesIO(raw_bytes))
                xl_img.width, xl_img.height = disp_w, disp_h
                ws.add_image(xl_img, f"A{r_idx}")
                ws.row_dimensions[r_idx].height = max(60, round(disp_h * 0.75) + 6)
            except Exception as e:
                print(f"[xlsx export] embedding cover failed for book {r['id']}: {e}")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="my-library-export.xlsx",
    )


@app.route("/api/import/xlsx", methods=["POST"])
def import_xlsx():
    file = request.files.get("file")
    if not file:
        return jsonify({"error": "No file uploaded."}), 400
    try:
        wb = load_workbook(io.BytesIO(file.read()))
    except Exception:
        return jsonify({"error": "Could not read the file as an Excel workbook."}), 400

    ws = wb.active
    rows_iter = ws.iter_rows(min_row=1, values_only=True)
    try:
        header = [str(h) if h is not None else "" for h in next(rows_iter)]
    except StopIteration:
        return jsonify({"error": "The Excel file is empty."}), 400

    cover_by_row = {}
    for img in ws._images:
        try:
            cover_by_row[img.anchor._from.row + 1] = img._data()
        except Exception:
            continue

    def cells_to_row_dict(cells, row_num):
        d = {header[i]: ("" if cells[i] is None else str(cells[i])) for i in range(min(len(header), len(cells)))}
        if row_num in cover_by_row:
            d["_coverBytes"] = cover_by_row[row_num]
        return d

    rows = (cells_to_row_dict(cells, row_num) for row_num, cells in enumerate(rows_iter, start=2))
    db = get_db()
    result = process_import_rows(db, rows)
    return jsonify(result)


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=False)
